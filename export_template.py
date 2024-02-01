import openpyxl 
import pymongo
import datetime as dt
import json
from azure.storage.blob import ContainerClient, BlobServiceClient, BlobClient
import os

engagetabname = "Engagement"
workingtabname = "Working"
cosmos_db_name = "sql-model-review"
collection_name = "kitt-collection"
output_container = "sql-model-review-template"


def connect_cusmosdb(filename: str, prompt_list: list):
    """select the files and date when export the findings"""
    try:
        CONNECTION_STRING = "mongodb://kitt-sql-review:hfM2p34q75uIxLRJtQfYNbZk8Y4RmAJMSlb5SRtXp6NfMhVQfzSevb7FcQ95MJdGqqX59bQQoHp3ACDbvHrHgw==@kitt-sql-review.mongo.cosmos.azure.com:10255/?ssl=true&replicaSet=globaldb&retrywrites=false&maxIdleTimeMS=120000&appName=@kitt-sql-review@"
        mglient = pymongo.MongoClient(CONNECTION_STRING)
        collection = mglient[cosmos_db_name][collection_name]
        select_prompts = {}
        for doc in collection.find({"queryId": filename}):
            for i in prompt_list:
                for j in doc["responses"]:
                    res = [j[i] for k in j.keys() if i == k ]
                    if len(res) > 0:
                        select_prompts.update({i : res[0]})

        return select_prompts
    except Exception as e:
        print(f"conencting to {collection_name} to extract the {filename} document failed, because {e}")


def update_engagement_dic(filename: str, prompt_list: list ):
    """Update the info with the file Name and number of findings exported"""
    data_orgi = json.load(open("engagementinfo.json"))
    data_orgi["File Name:"] = filename
    data_orgi["Number of Findings:"] = len(prompt_list)

    return data_orgi


def save_file_container(file_name: str):
    try:
        output_file = file_name + '.xlsx'
        blobService = BlobServiceClient.from_connection_string(conn_str = "DefaultEndpointsProtocol=https;AccountName=prodkittsqlmodel;AccountKey=dH2zsQDaDtA4fBgPQMCK1TaVM/F6HCQ56kG+6fHqddulxR8FAe6V/JPh8wKRjiwNmm1/8F1WPUVn+AStAxElrg==;EndpointSuffix=core.windows.net")
        blob_client = blobService.get_blob_client(output_container, output_file)
        with open(output_file, "rb") as data:
            blob_client.upload_blob(data, overwrite=True)

    except Exception as e:
        print(f"The {file_name} wasn't successfully exported because {e}")

def write_working_tab(filename: str, prompt_list: list):
    """Read the Excel template and write the relavent info to tabs"""
    wb = openpyxl.load_workbook("SQLModelReviewTemplate.xlsx")
    ws1 = wb.get_sheet_by_name(engagetabname)
    engage_info = update_engagement_dic(filename , prompt_list)
    ####update the info in the engagement tab
    ws1["C2"] = dt.date.today()
    for k, v in engage_info.items():
        for row in ws1.iter_rows(min_row=4, max_row=15, min_col=2, max_col=2):
            for cell in row:
                if cell.value == k:
                    ws1.cell(row=cell.row , column=3).value = v

    ####update the findings in the working tab
    ws2 = wb.get_sheet_by_name(workingtabname)
    findings = connect_cusmosdb(filename , prompt_list)  ##file reviewed and output findings
    finding_output_start_index = 12
    for q in findings.keys():
        if q == "Summarise the SQL query.":
            ws2.cell(row= 2, column=2).value = findings["Summarise the SQL query."]
        else:
            ws2.cell(row= finding_output_start_index, column=2).value = q
            ws2.cell(row= finding_output_start_index, column=3).value = findings[q]
            finding_output_start_index += 1
    
    output_file = filename + '.xlsx'
    wb.save(output_file)
    wb.close()



def main():
    file_reviewed = "create_function_dbo_getsizeband"
    findings_output = ["Summarise the SQL query.", "what are the inputs?", "what are the risks running this query?"]
    write_working_tab(file_reviewed, findings_output)
    save_file_container(file_reviewed)
    os.remove(file_reviewed+'.xlsx')

main()